"""
Comprehensive model comparison on Training_Data_v3/test.jsonl.
Evaluates all models SEQUENTIALLY (load → eval → unload → next).
Designed for H100 80 GB but works on any single GPU.

Models evaluated:
  qwen3b    — Qwen2.5-VL-3B   zero-shot (our base model, no fine-tuning)
  qwen7b    — Qwen2.5-VL-7B   zero-shot (larger baseline)
  internvl2 — InternVL2-2B    zero-shot (document-focused VLM)
  donut     — Donut DocVQA    zero-shot (document specialist, VQA-style)
  ours      — Qwen2.5-VL-3B + LoRA v3 fine-tuned (our model)

Metrics (where applicable):
  Class Accuracy  — predicted class == ground-truth class
  Position Acc.   — START/CONTINUATION prediction correct
  Field F1        — token-level F1 over non-null fields on START pages
  Split IoU       — Jaccard of START boundaries per packet

Usage
-----
# Full run (all 5 models, 500 stratified samples each)
python eval_all_models.py

# Select specific models
python eval_all_models.py --models qwen3b donut ours

# Quick smoke test
python eval_all_models.py --models qwen3b ours --max-samples 50

# Specify which v3 checkpoint to use for 'ours'
python eval_all_models.py --our-checkpoint model_output_v3/checkpoint-2500

Results saved to eval_results/comparison_<timestamp>/

Requirements:
  pip install openpyxl matplotlib numpy
"""

import argparse, json, os, re, time
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import torch

# ── Constants ─────────────────────────────────────────────────────────────────

CLASSES = [
    "Commercial Invoice", "House Bill of Lading", "Certificate of Origin",
    "Shipper's Letter of Instruction", "Dangerous Goods Declaration",
    "Verified Gross Mass", "House Airway Bill", "Packing List",
    "Customs Declaration", "Cargo Manifest", "Import/Export License",
    "Power of Attorney",
]
CLASS_SET = set(CLASSES)

START_FIELDS = [
    "shipper_name", "consignee_name", "document_date", "document_number",
    "country_of_origin", "country_of_destination", "description_of_goods",
    "license_number", "validity_start", "validity_end", "licensee_name",
]

FIELD_DISPLAY_NAMES = {
    "shipper_name":           "Shipper",
    "consignee_name":         "Consignee",
    "document_date":          "Date",
    "document_number":        "Doc. Number",
    "country_of_origin":      "Origin",
    "country_of_destination": "Destination",
    "description_of_goods":   "Goods Desc.",
    "license_number":         "License No.",
    "validity_start":         "Valid From",
    "validity_end":           "Valid Until",
    "licensee_name":          "Licensee",
}

# Same prompt used in Training_Data_v3 (no weight fields, with Rules)
PROMPT = (
    "Analyze this DHL logistics document page.\n\n"
    "Previous page: {prev}\n\n"
    "Output a single JSON line.\n\n"
    "If this is the START (first page of a new document):\n"
    '  {{"class": "...", "position": "START", "shipper_name": "...", '
    '"consignee_name": "...", "document_date": "...", "document_number": "...", '
    '"country_of_origin": "...", "country_of_destination": "...", '
    '"description_of_goods": "...", '
    '"license_number": "...", "validity_start": "...", '
    '"validity_end": "...", "licensee_name": "..."}}\n\n'
    "If this is a CONTINUATION (same document continues from previous page):\n"
    '  {{"class": "...", "position": "CONTINUATION"}}\n\n'
    "Document classes: " + ", ".join(CLASSES) + "\n\n"
    "Rules:\n"
    "- Output ONLY values that are clearly printed and readable on this page.\n"
    "- If a field is blank, empty, missing, or you cannot read it, output null.\n"
    "- Do NOT invent, guess, or fill in values from memory.\n"
    "- Use null for all fields not visible on this page."
)

# Donut classification question
DONUT_CLASS_QUESTION = (
    "What type of logistics document is this page? "
    "Answer with exactly one of: " + ", ".join(CLASSES) + "."
)
DONUT_POS_QUESTION = (
    "Is this the first page of a new document (answer START) "
    "or a continuation of a previous document (answer CONTINUATION)?"
)


# ── Model registry ─────────────────────────────────────────────────────────────

def _find_our_checkpoint(base_dir: Path) -> Path | None:
    """Auto-find the highest-step checkpoint in model_output_v3/."""
    v3 = base_dir / "model_output_v3"
    if not v3.exists():
        return None
    checkpoints = sorted(
        [d for d in v3.iterdir() if d.is_dir() and d.name.startswith("checkpoint-")],
        key=lambda d: int(d.name.split("-")[1]),
    )
    return checkpoints[-1] if checkpoints else None


def build_model_registry(base_dir: Path, our_checkpoint: str | None) -> dict:
    our_path = Path(our_checkpoint) if our_checkpoint else _find_our_checkpoint(base_dir)
    if our_path and not our_path.is_absolute():
        our_path = base_dir / our_path

    checkpoint_name = our_path.name if our_path else "NOT FOUND"

    return {
        "qwen3b": {
            "label":       "Qwen2.5-VL-3B  |  Zero-Shot  |  General VLM",
            "short_label": "Qwen-3B-ZS",
            "type":        "qwen",
            "path":        str(base_dir / "models" / "Qwen2.5-VL-3B-Instruct"),
            "metrics":     "all",
        },
        "qwen7b": {
            "label":       "Qwen2.5-VL-7B  |  Zero-Shot  |  General VLM (2.3x Larger)",
            "short_label": "Qwen-7B-ZS",
            "type":        "qwen",
            "path":        str(base_dir / "models" / "Qwen2.5-VL-7B-Instruct"),
            "metrics":     "all",
        },
        "internvl2": {
            "label":       "InternVL2-2B  |  Zero-Shot  |  Document-Focused VLM",
            "short_label": "InternVL2-ZS",
            "type":        "internvl2",
            "path":        "OpenGVLab/InternVL2-2B",
            "metrics":     "all",
        },
        "donut": {
            "label":       "Donut-DocVQA  |  Zero-Shot  |  Document Specialist",
            "short_label": "Donut-ZS",
            "type":        "donut",
            "path":        "naver-clova-ix/donut-base-finetuned-docvqa",
            "metrics":     "class_only",   # VQA approach: class + position, field F1 N/A
        },
        "ours": {
            "label":       f"Ours: Qwen2.5-VL-3B  |  Fine-Tuned v3  |  {checkpoint_name}",
            "short_label": "Ours-FT",
            "type":        "ours",
            "path":        str(our_path) if our_path else None,
            "base":        str(base_dir / "models" / "Qwen2.5-VL-3B-Instruct"),
            "metrics":     "all",
        },
    }


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Compare all models on test.jsonl")
    p.add_argument("--models", nargs="+",
                   choices=["qwen3b", "qwen7b", "internvl2", "donut", "ours"],
                   default=["qwen3b", "qwen7b", "internvl2", "donut", "ours"],
                   help="Which models to evaluate (default: all)")
    p.add_argument("--our-checkpoint", default=None,
                   help="Path to fine-tuned v3 checkpoint (default: auto-detect latest)")
    p.add_argument("--test-jsonl",  default="Training_Data_v3/test.jsonl")
    p.add_argument("--image-base",  default="Training_Data")
    p.add_argument("--max-samples", type=int, default=500,
                   help="Max samples per model, stratified (0=all)")
    p.add_argument("--max-new-tokens", type=int, default=220)
    p.add_argument("--resume",      action="store_true",
                   help="Skip models whose results file already exists")
    return p.parse_args()


# ── Data loading & sampling ────────────────────────────────────────────────────

def load_and_sample(jsonl_path: Path, max_n: int, seed: int = 42) -> list:
    import random
    with open(jsonl_path) as f:
        rows = [json.loads(l) for l in f]
    if max_n <= 0 or len(rows) <= max_n:
        return rows
    random.seed(seed)
    by_class = defaultdict(list)
    for r in rows:
        cls = json.loads(r["label"]).get("class", "Unknown")
        by_class[cls].append(r)
    per_cls = max(1, max_n // len(by_class))
    result  = []
    for cls, group in sorted(by_class.items()):
        random.shuffle(group)
        # Balance START/CONTINUATION within each class
        starts = [s for s in group if json.loads(s["label"]).get("position") == "START"]
        conts  = [s for s in group if json.loads(s["label"]).get("position") != "START"]
        half   = per_cls // 2
        result.extend(starts[:half])
        result.extend(conts[:per_cls - half])
    random.shuffle(result)
    return result[:max_n]


# ── Model loaders ──────────────────────────────────────────────────────────────

def load_qwen(model_path: str, device: str):
    from transformers import AutoProcessor
    try:
        from transformers import Qwen2_5_VLForConditionalGeneration as MC
    except ImportError:
        from transformers import AutoModelForVision2Seq as MC
    model = MC.from_pretrained(model_path, torch_dtype=torch.bfloat16, device_map=device)
    model.eval()
    processor = AutoProcessor.from_pretrained(model_path)
    print(f"  Loaded {Path(model_path).name}  ({sum(p.numel() for p in model.parameters())/1e9:.1f}B params)")
    return model, processor


def _load_qwen(model_path: str, device: str):
    return load_qwen(model_path, device)


def _load_internvl2(model_path: str, device: str):
    return load_internvl2(model_path, device)


def load_internvl2(model_path: str, device: str):
    from transformers import AutoModel, AutoTokenizer
    model = AutoModel.from_pretrained(
        model_path, torch_dtype=torch.bfloat16, device_map=device, trust_remote_code=True)
    model.eval()
    tokenizer = AutoTokenizer.from_pretrained(model_path, trust_remote_code=True)
    print(f"  Loaded InternVL2  ({sum(p.numel() for p in model.parameters())/1e9:.1f}B params)")
    return model, tokenizer


def load_donut(model_path: str, device: str):
    from transformers import DonutProcessor, VisionEncoderDecoderModel
    processor = DonutProcessor.from_pretrained(model_path)
    model = VisionEncoderDecoderModel.from_pretrained(
        model_path, torch_dtype=torch.bfloat16).to(device)
    model.eval()
    print(f"  Loaded Donut  ({sum(p.numel() for p in model.parameters())/1e6:.0f}M params)")
    return model, processor


def load_ours(lora_path: str, base_path: str, device: str):
    """Load fine-tuned model: base + LoRA adapters merged, plain transformers."""
    from transformers import AutoProcessor
    from peft import PeftModel
    try:
        from transformers import Qwen2_5_VLForConditionalGeneration as MC
    except ImportError:
        from transformers import AutoModelForVision2Seq as MC
    print(f"  Loading base model from {Path(base_path).name} ...")
    base  = MC.from_pretrained(base_path, torch_dtype=torch.bfloat16, device_map=device)
    print(f"  Loading LoRA from {Path(lora_path).name} and merging ...")
    peft  = PeftModel.from_pretrained(base, lora_path, torch_dtype=torch.bfloat16)
    model = peft.merge_and_unload()
    model.eval()
    processor = AutoProcessor.from_pretrained(lora_path)
    print(f"  Fine-tuned model ready  ({sum(p.numel() for p in model.parameters())/1e9:.1f}B params)")
    return model, processor


def unload(model):
    del model
    torch.cuda.empty_cache()


# ── Inference ──────────────────────────────────────────────────────────────────

_QWEN_TMPL_CACHE = None

def _qwen_prompt_text(processor, prev: str) -> str:
    global _QWEN_TMPL_CACHE
    if _QWEN_TMPL_CACHE is None:
        _QWEN_TMPL_CACHE = processor.apply_chat_template(
            [{"role": "user", "content": [
                {"type": "image"}, {"type": "text", "text": "PLACEHOLDER"}]}],
            tokenize=False, add_generation_prompt=True,
        )
    return _QWEN_TMPL_CACHE.replace("PLACEHOLDER", PROMPT.format(prev=prev))


def infer_qwen(model, processor, image, prev_label, max_new_tokens):
    text   = _qwen_prompt_text(processor, prev_label)
    inputs = processor(text=[text], images=[image], return_tensors="pt", padding=False)
    inputs = inputs.to(next(model.parameters()).device)
    with torch.inference_mode():
        out = model.generate(**inputs, max_new_tokens=max_new_tokens,
                             do_sample=False, use_cache=True)
    n = inputs["input_ids"].shape[1]
    return processor.decode(out[0][n:], skip_special_tokens=True).replace("<|im_end|>","").strip()


def _infer_qwen(model, processor, image, prev_label, max_new_tokens):
    return infer_qwen(model, processor, image, prev_label, max_new_tokens)


def _internvl2_pixel_values(image, device):
    import torchvision.transforms as T
    from torchvision.transforms.functional import InterpolationMode
    transform = T.Compose([
        T.Resize((448, 448), interpolation=InterpolationMode.BICUBIC),
        T.ToTensor(),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),
    ])
    img = image.convert("RGB")
    w, h = img.size
    tiles = []
    for row in range(2):
        for col in range(2):
            tiles.append(transform(img.crop((col*w//2, row*h//2, (col+1)*w//2, (row+1)*h//2))))
    tiles.append(transform(img))
    return torch.stack(tiles).to(torch.bfloat16).to(device)


def infer_internvl2(model, tokenizer, image, prev_label, max_new_tokens):
    pixel_values = _internvl2_pixel_values(image, next(model.parameters()).device)
    gen_cfg = dict(max_new_tokens=max_new_tokens, do_sample=False, use_cache=True)
    with torch.inference_mode():
        response = model.chat(tokenizer, pixel_values, PROMPT.format(prev=prev_label), gen_cfg)
    return response.strip() if isinstance(response, str) else str(response).strip()


def _infer_internvl2(model, tokenizer, image, prev_label, max_new_tokens):
    return infer_internvl2(model, tokenizer, image, prev_label, max_new_tokens)


def infer_donut(model, processor, image, device):
    """
    Donut VQA approach: two questions per page.
      Q1 -> document class
      Q2 -> START or CONTINUATION
    Returns raw text combining both answers for parse_response().
    """
    pixel_values = processor(image.convert("RGB"), return_tensors="pt").pixel_values
    pixel_values = pixel_values.to(torch.bfloat16).to(device)

    def ask(question):
        prompt = f"<s_docvqa><s_question>{question}</s_question><s_answer>"
        dec_ids = processor.tokenizer(prompt, add_special_tokens=False,
                                      return_tensors="pt").input_ids.to(device)
        with torch.inference_mode():
            out = model.generate(pixel_values, decoder_input_ids=dec_ids,
                                 max_length=100, early_stopping=True,
                                 pad_token_id=processor.tokenizer.pad_token_id,
                                 eos_token_id=processor.tokenizer.eos_token_id,
                                 use_cache=True, num_beams=1)
        seq = processor.tokenizer.decode(out[0], skip_special_tokens=False)
        # Extract answer between <s_answer> and </s_answer>
        m = re.search(r"<s_answer>(.*?)</s_answer>", seq, re.DOTALL)
        return m.group(1).strip() if m else seq.strip()

    class_answer = ask(DONUT_CLASS_QUESTION)
    pos_answer   = ask(DONUT_POS_QUESTION)

    # Match class answer to known classes (fuzzy)
    matched_cls = None
    for c in sorted(CLASS_SET, key=len, reverse=True):
        if c.lower() in class_answer.lower():
            matched_cls = c
            break

    # Match position
    matched_pos = None
    if re.search(r"\bCONTINUATION\b", pos_answer, re.IGNORECASE):
        matched_pos = "CONTINUATION"
    elif re.search(r"\bSTART\b", pos_answer, re.IGNORECASE):
        matched_pos = "START"
    elif "first" in pos_answer.lower() or "new" in pos_answer.lower():
        matched_pos = "START"
    elif "continuation" in pos_answer.lower() or "continue" in pos_answer.lower():
        matched_pos = "CONTINUATION"

    # Return as a pseudo-JSON string so parse_response() can handle it
    return json.dumps({"class": matched_cls, "position": matched_pos})


# ── Response parsing (same as dhl_app.py) ────────────────────────────────────

def parse_response(raw: str) -> dict:
    raw = raw.strip()
    try:
        obj = json.loads(raw)
        if isinstance(obj, dict):
            if obj.get("class") not in CLASS_SET: obj["class"] = None
            if obj.get("position") not in ("START", "CONTINUATION"): obj["position"] = None
            return obj
    except Exception: pass
    m = re.search(r"\{[^{}]+\}", raw, re.DOTALL)
    if m:
        try:
            obj = json.loads(m.group())
            if isinstance(obj, dict):
                if obj.get("class") not in CLASS_SET: obj["class"] = None
                if obj.get("position") not in ("START", "CONTINUATION"): obj["position"] = None
                return obj
        except Exception: pass
    found_cls = None
    for c in sorted(CLASS_SET, key=len, reverse=True):
        if c.lower() in raw.lower():
            found_cls = c; break
    found_pos = None
    if re.search(r"\bCONTINUATION\b", raw, re.IGNORECASE): found_pos = "CONTINUATION"
    elif re.search(r"\bSTART\b", raw, re.IGNORECASE): found_pos = "START"
    return {"class": found_cls, "position": found_pos}


# ── Field F1 helpers ──────────────────────────────────────────────────────────

def _token_f1(pred: str, true: str) -> float:
    p_tok = pred.lower().split(); t_tok = true.lower().split()
    if not p_tok or not t_tok: return 1.0 if p_tok == t_tok else 0.0
    common = set(p_tok) & set(t_tok)
    tp = sum(min(p_tok.count(t), t_tok.count(t)) for t in common)
    if not tp: return 0.0
    return 2 * (tp / len(p_tok)) * (tp / len(t_tok)) / ((tp / len(p_tok)) + (tp / len(t_tok)))


def field_metrics_detailed(pred: dict, true: dict):
    """Returns aggregate (tp, fp, fn) AND per-field breakdown dict."""
    total_tp = total_fp = total_fn = 0.0
    per_field = {}
    for f in START_FIELDS:
        gv = true.get(f); pv = pred.get(f)
        gh = gv is not None and str(gv).strip() != ""
        ph = pv is not None and str(pv).strip() != ""
        if gh and ph:
            f1 = _token_f1(str(pv).lower().strip(), str(gv).lower().strip())
            ftp, ffp, ffn = f1, 0.0, (1 - f1)
        elif gh:
            ftp, ffp, ffn = 0.0, 0.0, 1.0
        elif ph:
            ftp, ffp, ffn = 0.0, 1.0, 0.0
        else:
            ftp, ffp, ffn = 0.0, 0.0, 0.0   # TN — both null
        per_field[f] = {"tp": ftp, "fp": ffp, "fn": ffn}
        total_tp += ftp; total_fp += ffp; total_fn += ffn
    return total_tp, total_fp, total_fn, per_field


# ── Per-model evaluation loop ──────────────────────────────────────────────────

def evaluate_model(model_key: str, cfg: dict, samples: list,
                   image_base: Path, max_new_tokens: int,
                   out_dir: Path, resume: bool) -> dict:

    results_path = out_dir / f"{model_key}_results.json"
    if resume and results_path.exists():
        print(f"  Skipping {model_key} — results already exist at {results_path}")
        results = json.loads(results_path.read_text())
        return compute_metrics(results, cfg["label"], cfg["metrics"], cfg.get("short_label", ""))

    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"\n{'='*65}")
    print(f"  Evaluating: {cfg['label']}")
    print(f"  Samples   : {len(samples)}")
    print(f"{'='*65}")

    # ── Load model ────────────────────────────────────────────────────────────
    if cfg["type"] == "qwen" or cfg["type"] == "ours":
        if cfg["path"] is None:
            print("  ERROR: model path not found — skipping")
            return {"model": cfg["label"], "error": "path not found"}
        if not Path(cfg["path"]).exists():
            print(f"  ERROR: {cfg['path']} does not exist — skipping")
            return {"model": cfg["label"], "error": f"not found: {cfg['path']}"}

    if cfg["type"] == "qwen":
        model, processor = load_qwen(cfg["path"], device)
        infer_fn = lambda img, prev: infer_qwen(model, processor, img, prev, max_new_tokens)
    elif cfg["type"] == "internvl2":
        model, tokenizer = load_internvl2(cfg["path"], device)
        infer_fn = lambda img, prev: infer_internvl2(model, tokenizer, img, prev, max_new_tokens)
    elif cfg["type"] == "donut":
        model, processor = load_donut(cfg["path"], device)
        infer_fn = lambda img, prev: infer_donut(model, processor, img, device)
    elif cfg["type"] == "ours":
        model, processor = load_ours(cfg["path"], cfg["base"], device)
        infer_fn = lambda img, prev: infer_qwen(model, processor, img, prev, max_new_tokens)
    else:
        return {"model": cfg["label"], "error": "unknown type"}

    # ── Inference loop ────────────────────────────────────────────────────────
    results = []
    t_start = time.time()

    for i, sample in enumerate(samples):
        from PIL import Image
        img_path   = image_base / sample["image"]
        true_label = json.loads(sample["label"])
        prev_label = sample.get("prev_label", "none (first page of batch)")
        true_cls   = true_label.get("class")
        true_pos   = true_label.get("position")

        try:
            image = Image.open(img_path).convert("RGB")
        except Exception as e:
            print(f"  [SKIP] {img_path.name}: {e}")
            continue

        t0 = time.time()
        try:
            raw  = infer_fn(image, prev_label)
            pred = parse_response(raw)
        except Exception as e:
            print(f"  [ERR] {img_path.name}: {e}")
            pred = {}; raw = ""

        pred_cls = pred.get("class")
        pred_pos = pred.get("position")
        tp = fp = fn = 0.0
        per_field = {}
        if true_pos == "START":
            tp, fp, fn, per_field = field_metrics_detailed(pred, true_label)

        results.append({
            "image":         sample["image"],
            "true_class":    true_cls,
            "pred_class":    pred_cls,
            "true_position": true_pos,
            "pred_position": pred_pos,
            "class_correct": (pred_cls == true_cls),
            "pos_correct":   (pred_pos == true_pos),
            "is_start":      (true_pos == "START"),
            "field_tp": tp, "field_fp": fp, "field_fn": fn,
            "per_field":     per_field,
            "inference_s":   round(time.time() - t0, 2),
        })

        elapsed = time.time() - t_start
        eta     = elapsed / (i + 1) * (len(samples) - i - 1)
        print(f"  [{i+1:>4}/{len(samples)}] "
              f"cls={'ok' if pred_cls==true_cls else 'XX'}  "
              f"pos={'ok' if pred_pos==true_pos else 'XX'}  "
              f"{time.time()-t0:.1f}s  ETA {eta/60:.0f}m  "
              f"{Path(sample['image']).name}")

        # Save every 50 samples
        if (i + 1) % 50 == 0:
            results_path.write_text(json.dumps(results, indent=2))

    results_path.write_text(json.dumps(results, indent=2))

    # ── Unload ────────────────────────────────────────────────────────────────
    unload(model)

    return compute_metrics(results, cfg["label"], cfg["metrics"], cfg.get("short_label", ""))


# ── Metrics computation ────────────────────────────────────────────────────────

def _f1_from_tpfpfn(tp, fp, fn):
    prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    rec  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    return round(2 * prec * rec / (prec + rec) * 100, 1) if (prec + rec) > 0 else 0.0


def compute_metrics(results: list, label: str, metrics_type: str, short_label: str = "") -> dict:
    if not results:
        return {"model": label, "n": 0}

    # ── Overall accuracy ──────────────────────────────────────────────────────
    class_acc = sum(r["class_correct"] for r in results) / len(results) * 100
    pos_acc   = sum(r["pos_correct"]   for r in results) / len(results) * 100

    # ── Per-class breakdown ───────────────────────────────────────────────────
    per_class = defaultdict(lambda: {
        "total": 0, "cls_correct": 0,
        "start_total": 0, "start_correct": 0,
        "cont_total": 0,  "cont_correct": 0,
    })
    for r in results:
        cls = r["true_class"] or "Unknown"
        per_class[cls]["total"]       += 1
        per_class[cls]["cls_correct"] += int(r["class_correct"])
        if r["true_position"] == "START":
            per_class[cls]["start_total"]   += 1
            per_class[cls]["start_correct"] += int(r["pos_correct"])
        elif r["true_position"] == "CONTINUATION":
            per_class[cls]["cont_total"]    += 1
            per_class[cls]["cont_correct"]  += int(r["pos_correct"])

    per_class_acc = {
        cls: {
            "class_acc":   round(v["cls_correct"] / v["total"] * 100, 1) if v["total"] else 0,
            "start_acc":   round(v["start_correct"] / v["start_total"] * 100, 1) if v["start_total"] else None,
            "cont_acc":    round(v["cont_correct"]  / v["cont_total"]  * 100, 1) if v["cont_total"]  else None,
            "n":           v["total"],
        }
        for cls, v in sorted(per_class.items())
    }

    # ── Field F1 (aggregate + per-field) ─────────────────────────────────────
    starts = [r for r in results if r["is_start"]]
    if starts and metrics_type == "all":
        ttp = sum(r["field_tp"] for r in starts)
        tfp = sum(r["field_fp"] for r in starts)
        tfn = sum(r["field_fn"] for r in starts)
        overall_f1 = _f1_from_tpfpfn(ttp, tfp, tfn)

        per_field_f1 = {}
        for f in START_FIELDS:
            ftp = sum(r["per_field"].get(f, {}).get("tp", 0) for r in starts)
            ffp = sum(r["per_field"].get(f, {}).get("fp", 0) for r in starts)
            ffn = sum(r["per_field"].get(f, {}).get("fn", 0) for r in starts)
            per_field_f1[f] = _f1_from_tpfpfn(ftp, ffp, ffn)
    else:
        overall_f1   = None
        per_field_f1 = {f: None for f in START_FIELDS}

    # ── Split IoU ─────────────────────────────────────────────────────────────
    packets = defaultdict(list)
    for r in results:
        m = re.search(r"(packet_\d+)", r["image"])
        packets[m.group(1) if m else r["image"]].append(r)

    if metrics_type == "all":
        iou_scores = []
        for pages in packets.values():
            pages = sorted(pages, key=lambda r: int(re.search(r"_p(\d+)", r["image"]).group(1))
                           if re.search(r"_p(\d+)", r["image"]) else 0)
            true_s = set(j for j, r in enumerate(pages) if r["true_position"] == "START")
            pred_s = set(j for j, r in enumerate(pages) if r["pred_position"] == "START")
            u = len(true_s | pred_s)
            iou_scores.append(len(true_s & pred_s) / u if u else 1.0)
        split_iou = round(sum(iou_scores) / len(iou_scores) * 100, 1) if iou_scores else None
    else:
        split_iou = None

    # ── Per-class field F1 ────────────────────────────────────────────────────
    per_class_field_f1 = {}
    if metrics_type == "all":
        for cls in CLASSES:
            cls_starts = [r for r in results if r["is_start"] and r["true_class"] == cls]
            if cls_starts:
                tp = sum(r["field_tp"] for r in cls_starts)
                fp = sum(r["field_fp"] for r in cls_starts)
                fn = sum(r["field_fn"] for r in cls_starts)
                per_class_field_f1[cls] = _f1_from_tpfpfn(tp, fp, fn)
            else:
                per_class_field_f1[cls] = None
    else:
        per_class_field_f1 = {cls: None for cls in CLASSES}

    return {
        "model":              label,
        "short_label":        short_label if short_label else label[:15],
        "metrics_type":       metrics_type,
        "n":                  len(results),
        "class_accuracy":     round(class_acc, 1),
        "position_accuracy":  round(pos_acc, 1) if metrics_type == "all" else None,
        "field_f1":           overall_f1,
        "split_iou":          split_iou,
        "per_class":          dict(per_class_acc),
        "per_field_f1":       per_field_f1,
        "per_class_field_f1": per_class_field_f1,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(all_metrics: list, out_path: Path):
    """
    Write detailed results to Excel with 6 sheets + embedded charts.
    Requires: pip install openpyxl matplotlib
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Excel export skipped — run: pip install openpyxl")
        return

    # ── Colour palette (DHL brand) ────────────────────────────────────────────
    NAVY_HEX   = "002E5E"   # DHL dark navy — header background
    WHITE_HEX  = "FFFFFF"
    GREY_HEX   = "F0F4F8"   # label background
    GREEN_HEX  = "E8F5E9"   # our-model highlight
    DKGRN_HEX  = "1B5E20"   # our-model header text

    NAVY_FILL  = PatternFill("solid", fgColor=NAVY_HEX)
    GREY_FILL  = PatternFill("solid", fgColor=GREY_HEX)
    GREEN_FILL = PatternFill("solid", fgColor=GREEN_HEX)

    WHITE_BOLD = Font(color=WHITE_HEX, bold=True, name="Calibri", size=11)
    HEAD_FONT  = Font(bold=True,       name="Calibri", size=11)
    BODY_FONT  = Font(name="Calibri",  size=11)
    OUR_FONT   = Font(color=DKGRN_HEX, bold=True, name="Calibri", size=11)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _is_ours(m):
        return "Fine-Tuned" in m.get("model", "") or m.get("short_label", "") == "Ours-FT"

    def hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = NAVY_FILL; c.font = WHITE_BOLD
        c.alignment = CENTER; c.border = BORDER
        return c

    def label_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = GREY_FILL; c.font = HEAD_FONT
        c.alignment = LEFT; c.border = BORDER
        return c

    def val_cell(ws, row, col, value, pct=True, ours=False):
        if value is None:
            display = "N/A"
            c = ws.cell(row=row, column=col, value=display)
            c.font = BODY_FONT; c.alignment = CENTER; c.border = BORDER
            if ours:
                c.fill = GREEN_FILL
            return c
        if pct:
            display = f"{value:.1f}%"
        else:
            display = value
        c = ws.cell(row=row, column=col, value=display)
        c.alignment = CENTER; c.border = BORDER
        if ours:
            c.fill = GREEN_FILL
            c.font = OUR_FONT
        else:
            c.font = BODY_FONT
            # Colour-code percentages
            if pct and isinstance(value, (int, float)):
                if   value >= 90: c.fill = PatternFill("solid", fgColor="C6EFCE")
                elif value >= 75: c.fill = PatternFill("solid", fgColor="FFEB9C")
                elif value >= 60: c.fill = PatternFill("solid", fgColor="FFCC99")
                else:              c.fill = PatternFill("solid", fgColor="FFC7CE")
        return c

    def autofit_cols(ws, min_w=10, max_w=50):
        for col in ws.columns:
            best = min_w
            for cell in col:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, max_w))
            ws.column_dimensions[get_column_letter(col[0].column)].width = best

    N            = len(all_metrics)
    model_labels = [m["model"] for m in all_metrics]
    short_labels = [m.get("short_label", m["model"][:15]) for m in all_metrics]
    ours_flags   = [_is_ours(m) for m in all_metrics]

    wb = openpyxl.Workbook()

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 1: Summary
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    # Title
    title_cell = ws1.cell(row=1, column=1,
        value="DHL Document Intelligence — Model Comparison Results")
    title_cell.font      = Font(bold=True, name="Calibri", size=16, color=NAVY_HEX)
    title_cell.alignment = LEFT
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N + 1)
    ws1.row_dimensions[1].height = 30

    sub_cell = ws1.cell(row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sub_cell.font      = Font(italic=True, name="Calibri", size=10, color="555555")
    sub_cell.alignment = LEFT
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N + 1)

    # Blank row 3
    ws1.row_dimensions[3].height = 8

    # Header row 4
    hdr(ws1, 4, 1, "Model")
    for i, (lbl, sl, is_ours) in enumerate(zip(model_labels, short_labels, ours_flags)):
        c = ws1.cell(row=4, column=i + 2, value=lbl)
        c.alignment = CENTER; c.border = BORDER
        if is_ours:
            c.fill = PatternFill("solid", fgColor="1B5E20")
            c.font = Font(color=WHITE_HEX, bold=True, name="Calibri", size=11)
        else:
            c.fill = NAVY_FILL; c.font = WHITE_BOLD
    ws1.row_dimensions[4].height = 45

    SUMMARY_ROWS = [
        ("Samples Evaluated",    "n",                 False),
        ("Classification Acc.",  "class_accuracy",    True),
        ("Position Accuracy",    "position_accuracy", True),
        ("Field Extraction F1",  "field_f1",          True),
        ("Split IoU",            "split_iou",         True),
    ]
    for ri, (name, key, is_pct) in enumerate(SUMMARY_ROWS, start=5):
        label_cell(ws1, ri, 1, name)
        for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
            val_cell(ws1, ri, ci + 2, m.get(key), pct=is_pct, ours=is_ours)

    # Note row
    note_row = 5 + len(SUMMARY_ROWS) + 1
    nc = ws1.cell(row=note_row, column=1, value="* Higher is better for all metrics")
    nc.font = Font(italic=True, name="Calibri", size=9, color="888888")
    ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=N + 1)

    ws1.freeze_panes = ws1.cell(row=5, column=2)
    autofit_cols(ws1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 2: Per-Class Classification
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Class: Classification")
    ws2.row_dimensions[1].height = 45

    hdr(ws2, 1, 1, "Document Class")
    for i, (lbl, is_ours) in enumerate(zip(model_labels, ours_flags)):
        c = ws2.cell(row=1, column=i + 2, value=lbl)
        c.alignment = CENTER; c.border = BORDER
        c.fill = (PatternFill("solid", fgColor="1B5E20") if is_ours else NAVY_FILL)
        c.font = WHITE_BOLD

    for ri, cls in enumerate(CLASSES, start=2):
        label_cell(ws2, ri, 1, cls)
        for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
            v = m.get("per_class", {}).get(cls, {}).get("class_acc")
            val_cell(ws2, ri, ci + 2, v, pct=True, ours=is_ours)

    avg_row2 = len(CLASSES) + 2
    label_cell(ws2, avg_row2, 1, "AVERAGE")
    for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        val_cell(ws2, avg_row2, ci + 2, m.get("class_accuracy"), pct=True, ours=is_ours)

    ws2.freeze_panes = ws2.cell(row=2, column=2)
    autofit_cols(ws2)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 3: Per-Class Splitting
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Per-Class: Splitting")
    ws3.row_dimensions[1].height = 45
    ws3.row_dimensions[2].height = 30

    hdr(ws3, 1, 1, "Document Class")
    ws3.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    model_col_map3 = {}
    for m, is_ours in zip(all_metrics, ours_flags):
        ws3.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        mc = ws3.cell(row=1, column=col, value=m.get("short_label", m["model"][:15]))
        mc.alignment = CENTER; mc.border = BORDER
        mc.fill = (PatternFill("solid", fgColor="1B5E20") if is_ours else NAVY_FILL)
        mc.font = WHITE_BOLD
        for sub, sub_lbl in [(col, "START Acc"), (col + 1, "CONT Acc")]:
            c2 = ws3.cell(row=2, column=sub, value=sub_lbl)
            c2.fill = NAVY_FILL; c2.font = WHITE_BOLD
            c2.alignment = CENTER; c2.border = BORDER
        model_col_map3[m["model"]] = col
        col += 2

    for ri, cls in enumerate(CLASSES, start=3):
        label_cell(ws3, ri, 1, cls)
        for m, is_ours in zip(all_metrics, ours_flags):
            c = model_col_map3[m["model"]]
            pc = m.get("per_class", {}).get(cls, {})
            val_cell(ws3, ri, c,     pc.get("start_acc"), pct=True, ours=is_ours)
            val_cell(ws3, ri, c + 1, pc.get("cont_acc"),  pct=True, ours=is_ours)

    avg_row3 = len(CLASSES) + 3
    label_cell(ws3, avg_row3, 1, "AVERAGE")
    for m, is_ours in zip(all_metrics, ours_flags):
        c = model_col_map3[m["model"]]
        val_cell(ws3, avg_row3, c,     m.get("position_accuracy"), pct=True, ours=is_ours)
        val_cell(ws3, avg_row3, c + 1, m.get("position_accuracy"), pct=True, ours=is_ours)

    ws3.freeze_panes = ws3.cell(row=3, column=2)
    autofit_cols(ws3)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 4: Per-Class Field Extraction
    # ─────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Per-Class: Field Extraction")
    ws4.row_dimensions[1].height = 45

    hdr(ws4, 1, 1, "Document Class")
    for i, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        c = ws4.cell(row=1, column=i + 2, value=m["model"])
        c.alignment = CENTER; c.border = BORDER
        c.fill = (PatternFill("solid", fgColor="1B5E20") if is_ours else NAVY_FILL)
        c.font = WHITE_BOLD

    for ri, cls in enumerate(CLASSES, start=2):
        label_cell(ws4, ri, 1, cls)
        for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
            v = m.get("per_class_field_f1", {}).get(cls)
            val_cell(ws4, ri, ci + 2, v, pct=True, ours=is_ours)

    avg_row4 = len(CLASSES) + 2
    label_cell(ws4, avg_row4, 1, "AVERAGE")
    for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        val_cell(ws4, avg_row4, ci + 2, m.get("field_f1"), pct=True, ours=is_ours)

    note4_row = avg_row4 + 2
    nc4 = ws4.cell(row=note4_row, column=1,
        value="N/A for Donut (VQA approach cannot extract all fields)")
    nc4.font = Font(italic=True, name="Calibri", size=9, color="888888")
    ws4.merge_cells(start_row=note4_row, start_column=1, end_row=note4_row, end_column=N + 1)

    ws4.freeze_panes = ws4.cell(row=2, column=2)
    autofit_cols(ws4)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 5: Per-Field F1
    # ─────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Per-Field F1")
    ws5.row_dimensions[1].height = 45

    hdr(ws5, 1, 1, "Field")
    for i, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        c = ws5.cell(row=1, column=i + 2, value=m["model"])
        c.alignment = CENTER; c.border = BORDER
        c.fill = (PatternFill("solid", fgColor="1B5E20") if is_ours else NAVY_FILL)
        c.font = WHITE_BOLD

    for ri, f in enumerate(START_FIELDS, start=2):
        label_cell(ws5, ri, 1, FIELD_DISPLAY_NAMES.get(f, f))
        for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
            v = m.get("per_field_f1", {}).get(f)
            val_cell(ws5, ri, ci + 2, v, pct=True, ours=is_ours)

    avg_row5 = len(START_FIELDS) + 2
    label_cell(ws5, avg_row5, 1, "AVERAGE")
    for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        val_cell(ws5, avg_row5, ci + 2, m.get("field_f1"), pct=True, ours=is_ours)

    ws5.freeze_panes = ws5.cell(row=2, column=2)
    autofit_cols(ws5)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 6: Task Overview
    # ─────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Task Overview")
    ws6.row_dimensions[1].height = 50

    hdr(ws6, 1, 1, "Task")
    for i, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        c = ws6.cell(row=1, column=i + 2, value=m.get("short_label", m["model"][:15]))
        c.alignment = CENTER; c.border = BORDER
        c.fill = (PatternFill("solid", fgColor="1B5E20") if is_ours else NAVY_FILL)
        c.font = Font(color=WHITE_HEX, bold=True, name="Calibri", size=13)

    TASKS = [
        ("Document Classification (12 classes)",   "class_accuracy"),
        ("Bundle Splitting (START/CONT)",           "position_accuracy"),
        ("Field Extraction (11 fields)",            "field_f1"),
        ("Split IoU (Jaccard)",                     "split_iou"),
    ]
    for ri, (name, key) in enumerate(TASKS, start=2):
        c = ws6.cell(row=ri, column=1, value=name)
        c.fill = GREY_FILL
        c.font = Font(bold=True, name="Calibri", size=12)
        c.alignment = LEFT; c.border = BORDER
        ws6.row_dimensions[ri].height = 35
        for ci, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
            val_cell(ws6, ri, ci + 2, m.get(key), pct=True, ours=is_ours)

    ws6.freeze_panes = ws6.cell(row=2, column=2)
    autofit_cols(ws6)

    # ── Embed charts sheet (populated by generate_charts later if present) ────
    # Placeholder — charts sheet added after chart generation in main()

    wb.save(str(out_path))
    print(f"  Excel saved -> {out_path}")


# ── Chart generation ──────────────────────────────────────────────────────────

def generate_charts(all_metrics: list, out_dir: Path):
    """
    Generate 4 PNG charts and embed them in the Excel file's Charts sheet.
    Requires: pip install matplotlib numpy openpyxl
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        import numpy as np
    except ImportError:
        print("  Chart generation skipped — run: pip install matplotlib numpy")
        return []

    short_labels = [m.get("short_label", m["model"][:12]) for m in all_metrics]
    ours_flags   = ["Fine-Tuned" in m.get("model", "") or m.get("short_label", "") == "Ours-FT"
                    for m in all_metrics]

    DHL_NAVY  = "#002E5E"
    DHL_GREY  = "#8C9BA5"
    DHL_GREEN = "#2E7D32"
    THRESHOLD = 90

    bar_colors = [DHL_NAVY if f else DHL_GREY for f in ours_flags]

    def _bar_chart(ax, values, labels, title, colors, ylabel="%"):
        x = np.arange(len(labels))
        bars = ax.bar(x, [v if v is not None else 0 for v in values],
                      color=colors, width=0.6, edgecolor="white", linewidth=0.8)
        ax.axhline(THRESHOLD, color="#E53935", linestyle="--", linewidth=1.2,
                   label=f"{THRESHOLD}% threshold")
        ax.set_xticks(x)
        ax.set_xticklabels(labels, fontsize=9, rotation=15, ha="right")
        ax.set_ylim(0, 105)
        ax.set_ylabel(ylabel, fontsize=9)
        ax.set_title(title, fontsize=11, fontweight="bold", color=DHL_NAVY, pad=8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.yaxis.grid(True, linestyle="--", alpha=0.4)
        ax.set_axisbelow(True)
        for bar, v in zip(bars, values):
            if v is not None:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                        f"{v:.1f}", ha="center", va="bottom", fontsize=8,
                        color=DHL_NAVY, fontweight="bold")

    chart_paths = []

    # ── Chart 1: Overview (2x2) ───────────────────────────────────────────────
    fig1, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig1.suptitle("Model Comparison: DHL Document Intelligence Benchmark",
                  fontsize=14, fontweight="bold", color=DHL_NAVY, y=0.98)
    fig1.patch.set_facecolor("#FAFAFA")

    overview_data = [
        (axes[0, 0], [m.get("class_accuracy")    for m in all_metrics], "Classification Accuracy (%)"),
        (axes[0, 1], [m.get("position_accuracy")  for m in all_metrics], "Position Accuracy (%)"),
        (axes[1, 0], [m.get("field_f1")           for m in all_metrics], "Field Extraction F1 (%)"),
        (axes[1, 1], [m.get("split_iou")          for m in all_metrics], "Split IoU (%)"),
    ]
    for ax, vals, title in overview_data:
        _bar_chart(ax, vals, short_labels, title, bar_colors)

    our_patch  = mpatches.Patch(color=DHL_NAVY,  label="Our Model (Fine-Tuned)")
    base_patch = mpatches.Patch(color=DHL_GREY,  label="Baseline (Zero-Shot)")
    thr_line   = mpatches.Patch(color="#E53935", label=f"{THRESHOLD}% Threshold")
    fig1.legend(handles=[our_patch, base_patch, thr_line],
                loc="lower center", ncol=3, fontsize=9,
                framealpha=0.9, bbox_to_anchor=(0.5, 0.01))

    plt.tight_layout(rect=[0, 0.05, 1, 0.97])
    p1 = out_dir / "comparison_overview.png"
    fig1.savefig(str(p1), dpi=150, bbox_inches="tight")
    plt.close(fig1)
    chart_paths.append(p1)
    print(f"  Chart -> {p1}")

    # ── Chart 2: Radar ────────────────────────────────────────────────────────
    categories = ["Classification", "Position", "Field F1", "Split IoU"]
    keys       = ["class_accuracy", "position_accuracy", "field_f1", "split_iou"]
    n_cat      = len(categories)
    angles     = np.linspace(0, 2 * np.pi, n_cat, endpoint=False).tolist()
    angles    += angles[:1]

    fig2, ax2 = plt.subplots(figsize=(10, 10), subplot_kw={"projection": "polar"})
    fig2.patch.set_facecolor("#FAFAFA")
    ax2.set_facecolor("#F5F7FA")
    ax2.set_theta_offset(np.pi / 2)
    ax2.set_theta_direction(-1)
    ax2.set_xticks(angles[:-1])
    ax2.set_xticklabels(categories, fontsize=11, fontweight="bold", color=DHL_NAVY)
    ax2.set_ylim(0, 100)
    ax2.set_yticks([25, 50, 75, 100])
    ax2.set_yticklabels(["25%", "50%", "75%", "100%"], fontsize=8, color="#888888")
    ax2.yaxis.grid(True, linestyle="--", alpha=0.4)
    ax2.xaxis.grid(True, linestyle="-", alpha=0.2)

    palette = plt.cm.tab10.colors
    handles_radar = []
    for idx, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        vals = [m.get(k) or 0 for k in keys]
        vals += vals[:1]
        color = DHL_NAVY if is_ours else palette[idx % len(palette)]
        lw    = 2.5 if is_ours else 1.5
        ls    = "-" if is_ours else "--"
        alpha = 0.25 if is_ours else 0.0
        ax2.plot(angles, vals, color=color, linewidth=lw, linestyle=ls,
                 label=m.get("short_label", m["model"][:15]))
        if is_ours:
            ax2.fill(angles, vals, color=color, alpha=alpha)
        handles_radar.append(
            mpatches.Patch(color=color, label=m.get("short_label", m["model"][:15]))
        )

    ax2.set_title("Radar Comparison — All Models", fontsize=14, fontweight="bold",
                  color=DHL_NAVY, pad=20)
    ax2.legend(handles=handles_radar, loc="lower right", fontsize=9,
               bbox_to_anchor=(1.25, -0.05), framealpha=0.9)

    p2 = out_dir / "radar_comparison.png"
    fig2.savefig(str(p2), dpi=150, bbox_inches="tight")
    plt.close(fig2)
    chart_paths.append(p2)
    print(f"  Chart -> {p2}")

    # ── Chart 3: Per-Class Accuracy (horizontal grouped bar) ─────────────────
    fig3, ax3 = plt.subplots(figsize=(14, 10))
    fig3.patch.set_facecolor("#FAFAFA")
    ax3.set_facecolor("#F5F7FA")

    y      = np.arange(len(CLASSES))
    n_m    = len(all_metrics)
    h      = 0.8 / n_m
    offset = -(0.4 - h / 2)

    for idx, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        vals   = [m.get("per_class", {}).get(cls, {}).get("class_acc") or 0 for cls in CLASSES]
        color  = DHL_NAVY if is_ours else palette[idx % len(palette)]
        ax3.barh(y + offset + idx * h, vals, height=h * 0.9, color=color,
                 label=m.get("short_label", m["model"][:15]))

    ax3.axvline(THRESHOLD, color="#E53935", linestyle="--", linewidth=1.2,
                label=f"{THRESHOLD}% threshold")
    ax3.set_yticks(y)
    ax3.set_yticklabels(CLASSES, fontsize=9)
    ax3.set_xlim(0, 108)
    ax3.set_xlabel("Classification Accuracy (%)", fontsize=10)
    ax3.set_title("Per-Class Classification Accuracy — All Models",
                  fontsize=13, fontweight="bold", color=DHL_NAVY, pad=10)
    ax3.spines["top"].set_visible(False)
    ax3.spines["right"].set_visible(False)
    ax3.xaxis.grid(True, linestyle="--", alpha=0.4)
    ax3.set_axisbelow(True)
    ax3.legend(loc="lower right", fontsize=9, framealpha=0.9)

    plt.tight_layout()
    p3 = out_dir / "per_class_accuracy.png"
    fig3.savefig(str(p3), dpi=150, bbox_inches="tight")
    plt.close(fig3)
    chart_paths.append(p3)
    print(f"  Chart -> {p3}")

    # ── Chart 4: Per-Field F1 (grouped vertical bar) ─────────────────────────
    field_labels = [FIELD_DISPLAY_NAMES.get(f, f) for f in START_FIELDS]
    fig4, ax4 = plt.subplots(figsize=(14, 7))
    fig4.patch.set_facecolor("#FAFAFA")
    ax4.set_facecolor("#F5F7FA")

    x4     = np.arange(len(START_FIELDS))
    h4     = 0.8 / n_m
    off4   = -(0.4 - h4 / 2)

    for idx, (m, is_ours) in enumerate(zip(all_metrics, ours_flags)):
        vals   = [m.get("per_field_f1", {}).get(f) or 0 for f in START_FIELDS]
        color  = DHL_NAVY if is_ours else palette[idx % len(palette)]
        ax4.bar(x4 + off4 + idx * h4, vals, width=h4 * 0.9, color=color,
                label=m.get("short_label", m["model"][:15]))

    ax4.axhline(THRESHOLD, color="#E53935", linestyle="--", linewidth=1.2,
                label=f"{THRESHOLD}% threshold")
    ax4.set_xticks(x4)
    ax4.set_xticklabels(field_labels, fontsize=9, rotation=20, ha="right")
    ax4.set_ylim(0, 110)
    ax4.set_ylabel("Token-Level F1 (%)", fontsize=10)
    ax4.set_title("Per-Field Extraction F1 — All Models",
                  fontsize=13, fontweight="bold", color=DHL_NAVY, pad=10)
    ax4.spines["top"].set_visible(False)
    ax4.spines["right"].set_visible(False)
    ax4.yaxis.grid(True, linestyle="--", alpha=0.4)
    ax4.set_axisbelow(True)
    ax4.legend(loc="upper right", fontsize=9, framealpha=0.9)

    plt.tight_layout()
    p4 = out_dir / "per_field_f1.png"
    fig4.savefig(str(p4), dpi=150, bbox_inches="tight")
    plt.close(fig4)
    chart_paths.append(p4)
    print(f"  Chart -> {p4}")

    # ── Embed charts into Excel (Charts sheet) ────────────────────────────────
    excel_path = out_dir / "comparison_results.xlsx"
    if excel_path.exists():
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            wb2 = openpyxl.load_workbook(str(excel_path))
            if "Charts" in wb2.sheetnames:
                del wb2["Charts"]
            wsc = wb2.create_sheet("Charts")

            # 2x2 grid: anchor positions (row, col) in Excel cell notation
            anchors = ["B2", "P2", "B32", "P32"]
            for anchor, chart_path in zip(anchors, chart_paths):
                if chart_path.exists():
                    img = XLImage(str(chart_path))
                    img.width  = 500
                    img.height = 360
                    wsc.add_image(img, anchor)

            wsc.sheet_view.showGridLines = False
            title_c = wsc.cell(row=1, column=2,
                value="DHL Document Intelligence — Visual Comparison")
            title_c.font = openpyxl.styles.Font(
                bold=True, name="Calibri", size=14, color="002E5E")

            wb2.save(str(excel_path))
            print(f"  Charts embedded -> {excel_path} (Charts sheet)")
        except Exception as e:
            print(f"  Warning: could not embed charts in Excel: {e}")

    return chart_paths


# ── Comparison table ───────────────────────────────────────────────────────────

def print_comparison_table(all_metrics: list):
    print(f"\n{'='*80}")
    print("  COMPARISON TABLE")
    print(f"{'='*80}")
    header = (f"  {'Model (short)':<18} {'Samples':>8} "
              f"{'Cls Acc':>9} {'Pos Acc':>9} {'Field F1':>9} {'Split IoU':>10}")
    print(header)
    print(f"  {'-'*74}")
    for m in all_metrics:
        if "error" in m:
            lbl = m.get("short_label", m.get("model", "?"))
            print(f"  {lbl:<18}  ERROR: {m['error']}")
            continue
        def fmt(v):
            return f"{v:.1f}%" if v is not None else "   N/A"
        sl = m.get("short_label", m.get("model", "?")[:18])
        print(
            f"  {sl:<18}"
            f"  {m.get('n', 0):>8}"
            f"  {fmt(m.get('class_accuracy')):>9}"
            f"  {fmt(m.get('position_accuracy')):>9}"
            f"  {fmt(m.get('field_f1')):>9}"
            f"  {fmt(m.get('split_iou')):>9}"
        )
    print(f"{'='*80}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args     = parse_args()
    base     = Path(__file__).parent
    registry = build_model_registry(base, args.our_checkpoint)

    # Output dir with timestamp
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = base / "eval_results" / f"comparison_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load test samples (same set for all models)
    test_path  = base / args.test_jsonl
    image_base = base / args.image_base
    print(f"Loading test data from: {test_path}")
    all_samples = load_and_sample(test_path, args.max_samples)
    print(f"Using {len(all_samples)} samples (same for all models)\n")

    all_metrics   = []
    total_t_start = time.time()

    for model_key in args.models:
        if model_key not in registry:
            print(f"Unknown model key: {model_key} — skipping")
            continue

        cfg     = registry[model_key]
        metrics = evaluate_model(
            model_key, cfg, all_samples, image_base,
            args.max_new_tokens, out_dir, args.resume,
        )
        all_metrics.append(metrics)

        # Save after every model — JSON + Excel both updated continuously
        (out_dir / "comparison.json").write_text(json.dumps(all_metrics, indent=2))
        export_to_excel(all_metrics, out_dir / "comparison_results.xlsx")

    # Generate charts and embed in Excel
    generate_charts(all_metrics, out_dir)

    total_elapsed = time.time() - total_t_start
    print_comparison_table(all_metrics)

    excel_rel  = out_dir / "comparison_results.xlsx"
    chart_rels = [
        out_dir / "comparison_overview.png",
        out_dir / "radar_comparison.png",
        out_dir / "per_class_accuracy.png",
        out_dir / "per_field_f1.png",
    ]

    print(f"Total evaluation time: {total_elapsed/60:.0f} min")
    print(f"Results saved to:  {out_dir}")
    print(f"Excel:             {excel_rel}")
    print(f"Charts:")
    for cp in chart_rels:
        print(f"  {cp}")


if __name__ == "__main__":
    main()
